"""
Data Exporter for exporting test results in various formats
"""

import logging
import json
import csv
from typing import Dict, List, Any, Optional
from datetime import datetime
import os

logger = logging.getLogger(__name__)

class DataExporter:
    """Exports test data in various formats"""
    
    def __init__(self, config):
        """Initialize data exporter"""
        self.config = config
        
    def export_results(self, results: Dict[str, Any], output_dir: str, 
                      formats: List[str] = None) -> Dict[str, str]:
        """Export results in specified formats"""
        
        if formats is None:
            formats = ['json', 'csv']
        
        os.makedirs(output_dir, exist_ok=True)
        
        exported_files = {}
        
        for format_type in formats:
            try:
                if format_type == 'json':
                    file_path = self._export_json(results, output_dir)
                elif format_type == 'csv':
                    file_path = self._export_csv(results, output_dir)
                elif format_type == 'excel':
                    file_path = self._export_excel(results, output_dir)
                else:
                    logger.warning(f"Unsupported export format: {format_type}")
                    continue
                
                exported_files[format_type] = file_path
                logger.info(f"Exported {format_type} to {file_path}")
                
            except Exception as e:
                logger.error(f"Failed to export {format_type}: {e}")
        
        return exported_files
    
    def _export_json(self, results: Dict[str, Any], output_dir: str) -> str:
        """Export results as JSON"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_results_{timestamp}.json"
        file_path = os.path.join(output_dir, filename)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        
        return file_path
    
    def _export_csv(self, results: Dict[str, Any], output_dir: str) -> str:
        """Export results as CSV"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_results_{timestamp}.csv"
        file_path = os.path.join(output_dir, filename)
        
        # Flatten results for CSV export
        flattened_data = self._flatten_results(results)
        
        if flattened_data:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=flattened_data[0].keys())
                writer.writeheader()
                writer.writerows(flattened_data)
        
        return file_path
    
    def _export_excel(self, results: Dict[str, Any], output_dir: str) -> str:
        """Export results as Excel (requires openpyxl)"""
        
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl not installed. Cannot export to Excel.")
            raise
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_results_{timestamp}.xlsx"
        file_path = os.path.join(output_dir, filename)
        
        wb = Workbook()
        
        # Create summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary data
        summary_data = self._prepare_summary_data(results)
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)
        
        # Create detailed results sheet
        ws_details = wb.create_sheet("Detailed Results")
        flattened_data = self._flatten_results(results)
        
        if flattened_data:
            # Add headers
            headers = list(flattened_data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                ws_details.cell(row=1, column=col_idx, value=header)
            
            # Add data
            for row_idx, row_data in enumerate(flattened_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_details.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
        
        wb.save(file_path)
        return file_path
    
    def _flatten_results(self, results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Flatten nested results for tabular export"""
        
        flattened = []
        
        if 'test_results' in results:
            for model_name, model_data in results['test_results'].items():
                if isinstance(model_data, dict):
                    for test_type, test_data in model_data.items():
                        if isinstance(test_data, dict):
                            row = {
                                'session_id': results.get('session_id', ''),
                                'timestamp': results.get('timestamp', ''),
                                'model': model_name,
                                'test_type': test_type,
                                'score': test_data.get('score', 0),
                                'status': 'passed' if test_data.get('score', 0) >= 0.7 else 'failed',
                                'details': str(test_data.get('details', ''))
                            }
                            flattened.append(row)
        
        return flattened
    
    def _prepare_summary_data(self, results: Dict[str, Any]) -> List[List[Any]]:
        """Prepare summary data for Excel export"""
        
        summary_data = [
            ['Test Session Summary'],
            ['Session ID', results.get('session_id', 'N/A')],
            ['Timestamp', results.get('timestamp', 'N/A')],
            ['Scope', results.get('scope', 'N/A')],
            ['Models Tested', ', '.join(results.get('models_tested', []))],
            [''],
            ['Model Performance Summary']
        ]
        
        if 'test_results' in results:
            summary_data.append(['Model', 'Average Score', 'Tests Passed', 'Tests Failed'])
            
            for model_name, model_data in results['test_results'].items():
                if isinstance(model_data, dict):
                    scores = []
                    passed = 0
                    failed = 0
                    
                    for test_data in model_data.values():
                        if isinstance(test_data, dict) and 'score' in test_data:
                            score = test_data['score']
                            scores.append(score)
                            if score >= 0.7:
                                passed += 1
                            else:
                                failed += 1
                    
                    avg_score = sum(scores) / len(scores) if scores else 0.0
                    summary_data.append([model_name, f"{avg_score:.3f}", passed, failed])
        
        return summary_data
    
    def export_comparison_data(self, results_list: List[Dict[str, Any]], 
                              output_dir: str) -> str:
        """Export comparison data across multiple test sessions"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"comparison_results_{timestamp}.csv"
        file_path = os.path.join(output_dir, filename)
        
        comparison_data = []
        
        for results in results_list:
            session_id = results.get('session_id', 'unknown')
            timestamp = results.get('timestamp', 'unknown')
            
            if 'test_results' in results:
                for model_name, model_data in results['test_results'].items():
                    if isinstance(model_data, dict):
                        for test_type, test_data in model_data.items():
                            if isinstance(test_data, dict):
                                row = {
                                    'session_id': session_id,
                                    'timestamp': timestamp,
                                    'model': model_name,
                                    'test_type': test_type,
                                    'score': test_data.get('score', 0),
                                    'status': 'passed' if test_data.get('score', 0) >= 0.7 else 'failed'
                                }
                                comparison_data.append(row)
        
        if comparison_data:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=comparison_data[0].keys())
                writer.writeheader()
                writer.writerows(comparison_data)
        
        logger.info(f"Comparison data exported to {file_path}")
        return file_path
    
    def export_trend_analysis(self, historical_results: List[Dict[str, Any]], 
                             output_dir: str) -> str:
        """Export trend analysis data"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trend_analysis_{timestamp}.json"
        file_path = os.path.join(output_dir, filename)
        
        trend_analysis = {
            'analysis_timestamp': datetime.now().isoformat(),
            'data_points': len(historical_results),
            'trends': self._calculate_trends(historical_results),
            'summary': self._generate_trend_summary(historical_results)
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(trend_analysis, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Trend analysis exported to {file_path}")
        return file_path
    
    def _calculate_trends(self, historical_results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate performance trends from historical data"""
        
        trends = {}
        
        # Group by model and test type
        model_test_scores = {}
        
        for results in historical_results:
            if 'test_results' in results:
                for model_name, model_data in results['test_results'].items():
                    if model_name not in model_test_scores:
                        model_test_scores[model_name] = {}
                    
                    if isinstance(model_data, dict):
                        for test_type, test_data in model_data.items():
                            if test_type not in model_test_scores[model_name]:
                                model_test_scores[model_name][test_type] = []
                            
                            if isinstance(test_data, dict) and 'score' in test_data:
                                model_test_scores[model_name][test_type].append(test_data['score'])
        
        # Calculate trends for each model-test combination
        for model_name, test_types in model_test_scores.items():
            trends[model_name] = {}
            
            for test_type, scores in test_types.items():
                if len(scores) >= 2:
                    # Simple trend calculation
                    first_half = scores[:len(scores)//2]
                    second_half = scores[len(scores)//2:]
                    
                    first_avg = sum(first_half) / len(first_half)
                    second_avg = sum(second_half) / len(second_half)
                    
                    trend_direction = 'improving' if second_avg > first_avg else 'declining' if second_avg < first_avg else 'stable'
                    trend_magnitude = abs(second_avg - first_avg)
                    
                    trends[model_name][test_type] = {
                        'direction': trend_direction,
                        'magnitude': trend_magnitude,
                        'first_period_avg': first_avg,
                        'second_period_avg': second_avg,
                        'data_points': len(scores)
                    }
        
        return trends
    
    def _generate_trend_summary(self, historical_results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate summary of trends"""
        
        summary = {
            'total_sessions': len(historical_results),
            'date_range': self._get_date_range(historical_results),
            'overall_performance': self._calculate_overall_performance_trend(historical_results)
        }
        
        return summary
    
    def _get_date_range(self, historical_results: List[Dict[str, Any]]) -> Dict[str, str]:
        """Get date range of historical results"""
        
        timestamps = []
        for results in historical_results:
            if 'timestamp' in results:
                timestamps.append(results['timestamp'])
        
        if timestamps:
            timestamps.sort()
            return {
                'earliest': timestamps[0],
                'latest': timestamps[-1]
            }
        
        return {'earliest': 'unknown', 'latest': 'unknown'}
    
    def _calculate_overall_performance_trend(self, historical_results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate overall performance trend"""
        
        session_averages = []
        
        for results in historical_results:
            if 'test_results' in results:
                all_scores = []
                
                for model_data in results['test_results'].values():
                    if isinstance(model_data, dict):
                        for test_data in model_data.values():
                            if isinstance(test_data, dict) and 'score' in test_data:
                                all_scores.append(test_data['score'])
                
                if all_scores:
                    session_avg = sum(all_scores) / len(all_scores)
                    session_averages.append(session_avg)
        
        if len(session_averages) >= 2:
            first_half = session_averages[:len(session_averages)//2]
            second_half = session_averages[len(session_averages)//2:]
            
            first_avg = sum(first_half) / len(first_half)
            second_avg = sum(second_half) / len(second_half)
            
            return {
                'trend': 'improving' if second_avg > first_avg else 'declining' if second_avg < first_avg else 'stable',
                'change': second_avg - first_avg,
                'first_period_avg': first_avg,
                'second_period_avg': second_avg
            }
        
        return {'trend': 'insufficient_data', 'change': 0.0}

